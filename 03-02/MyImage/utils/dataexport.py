import openpyxl
from docx import Document
from reportlab.pdfgen import canvas

import datetime  # 导入日期时间库
import subprocess
import reportlab.pdfbase.ttfonts  # 导入reportlab的注册字体

reportlab.pdfbase.pdfmetrics.registerFont(
    reportlab.pdfbase.ttfonts.TTFont('song', 'C:\WINDOWS\Fonts\simsun.ttc'))  # 注册字体

from reportlab.pdfgen import canvas
from reportlab.lib.units import inch  # 导入单位英寸

import os
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

'''
{
    1: ['1.1', '小牛', '学习python'], 
    2: ['1.1', '小虎', '吃鸡'], 
    3: ['12.12', '老牛', '吃鸡'], 
    4: ['6.6', '小狗', '嘻嘻有'], 
    5: ['6.7', '小猫', '喝水'], 
    6: ['6.9', '小鸡鸡', '油水']
}

'''


class Export:
    '导出文件'
    def __init__(self, data: dict):
        self.data = data

    def get_absname(self, file_name):
        '生成文件的绝对路径'
        return os.path.join(BASE_DIR, 'db', file_name)

    def trans_data(self):
        '把字典数据转换为字符串，追加到列表中'
        memo_list = []
        for k, v in self.data.items():
            memo_thing = str(k).ljust(10,' ')
            for item in v:
                memo_thing += item.ljust(10, ' ')
            memo_thing += '\n'
            memo_list.append(memo_thing)

        return memo_list

    def export_txt(self, file_name='LH-Memo.txt'):
        '导出txt格式'
        with open(self.get_absname(file_name), 'w', encoding='utf-8') as f:
            for i in self.trans_data():
                f.write(i)

    def export_pdf(self, file_name='LH-Memo.pdf'):
        '导出pdf格式'
        now = datetime.datetime.today()
        date = now.strftime("%h %d %Y %H:%M:%S")  # 设定日期格式
        c = canvas.Canvas(self.get_absname(file_name))
        c.setFont('song', 10)  # 设置字体字号
        textobject = c.beginText()  # 定义开始
        textobject.setTextOrigin(inch, 11 * inch)  # 定义位置
        textobject.textLines('''LH-memo: %s ''' % date)  # 输出标题

        for line in self.trans_data():  # 通过循环的方式一行一行写入文件
            line = ' '.join(line)  # 拼接成字符串
            textobject.textLine(line.strip())  # 写入文件

        c.drawText(textobject)
        c.showPage()
        c.save()

    def export_excle(self, file_name='LH-Memo.xlsx'):
        '导出excel格式'
        wb = openpyxl.Workbook()
        sh = wb.active
        row = 2
        sh['A1'],sh['B1'] = '文件名', '大小'

        for item in self.data:
            sh.cell(row=row, column=1).value = item[0]
            sh.cell(row=row, column=2).value = item[1]
            row += 1
        wb.save(self.get_absname(file_name))

    def export_word(self,file_name='LH-Memo.docx'):
        '导出word格式'
        word = Document()
        for row in self.trans_data():
            word.add_paragraph(row)

        word.save(self.get_absname(file_name))
            


def main():
    dic = {
    1: ['1.1', '小牛', '学习python'],
    2: ['1.1', '小虎', '吃鸡'],
    3: ['12.12', '老牛', '吃鸡'],
    4: ['6.6', '小狗', '嘻嘻有'],
    5: ['6.7', '小猫', '喝水'],
    6: ['6.9', '小鸡鸡', '油水']
}
    ex = Export(dic)
    ex.export_excle()
    ex.export_pdf()
    ex.export_txt()
    ex.export_word()

if __name__ == '__main__':
    main()